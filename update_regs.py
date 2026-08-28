#!/usr/bin/env python3
"""
update_regs.py — Reads ALL Numbers/Excel files in the Santhosha Data folder,
extracts registration counts, and updates the dashboard HTML with:
  1. LIVE_REGS    — upcoming batch-level counts (reg badges on programme cards)
  2. CENTRE_DATA  — annual IE/BSP/Shoonya/Samyama totals (Centre Insights page)
  3. MONTHLY_DATA — month-by-month breakdown for trend charts (where available)

Run this script from Terminal whenever any file in Santhosha Data is updated:
    python3 update_regs.py

Requires: macOS (uses osascript to export Numbers → CSV)
          pip install openpyxl  (for reading .xlsx directly)
"""

import subprocess, csv, os, re, json, sys, tempfile, glob, datetime, urllib.request

# openpyxl — for reading .xlsx directly (no AppleScript needed)
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE          = os.path.dirname(os.path.abspath(__file__))
HTML          = os.path.join(BASE, "isha_multi_centre_dashboard_v2.html")
SANTHOSHA_DIR = os.path.join(BASE, "Santhosha Data")
CURRENT_YEAR  = str(datetime.date.today().year)

# ── Normalisation patterns ────────────────────────────────────────────────────
CENTRE_PATTERNS = [
    (r'sadhguru sannidhi',                                    'sadhguru sannidhi'),
    (r'isha yoga cent(?:er|re)|iyc|coimbatore|velliangiri',   'isha yoga center'),
    (r'electronic city',                                      'electronic'),
    (r'kanakapura',                                           'kanakapura'),
    (r'sarjapur|sargapur',                                    'sarjapur'),
    (r'hsr layout|hsr',                                       'hsr'),
    (r'marathahalli|marathali',                               'marathahalli'),
    (r'malleswaram|malleshwaram',                             'malleswaram'),
    (r'vijayanagar|vijayanagara',                             'vijayanagar'),
    (r'jayanagar|jaynagar',                                   'jayanagar'),
    (r'banaswadi|banasawadi',                                 'banaswadi'),
    (r'hebbal|hebbala',                                       'hebbal'),
    (r'indiranagar|indira nagar',                             'indiranagar'),
    (r'mysuru|mysore',                                        'mysuru'),
    (r'hubballi|hubbali|hubli',                               'hubbali'),
    (r'ballari|bellary',                                      'ballari'),
    (r'belagavi|belgaum',                                     'belagavi'),
    (r'koramangala',                                          'koramangala'),
    (r'chikkaballapur',                                       'chikkaballapur'),
    # Sub-centres (primarily Monthly Satsang, parsed from Pivot Event xlsx)
    (r'girinagar',                                            'girinagar'),
    (r'yelahanka',                                            'yelahanka'),
    (r'chandapura',                                           'chandapura'),
    (r'begur',                                                'begur'),
    (r'budigere cross',                                       'budigere cross'),  # full name first
    (r'budigere',                                             'budigere cross'),  # short alias → same key
    (r'mangalore|mangaluru',                                  'mangalore'),
    (r'kengeri',                                              'kengeri'),
    (r'peenya',                                               'peenya'),
    (r'bg road|bannerghatta road',                            'bg road'),
    (r'tumkur',                                               'tumkur'),
    (r'whitefield',                                           'whitefield'),
    (r'singasandra',                                          'singasandra'),
    (r'kodagu|coorg',                                         'kodagu'),
]

PROG_PATTERNS = [
    (r'inner engineering|isha yoga|\bie\b',                          'inner engineering'),
    (r'bhava spandana|\bbsp\b',                                      'bhava spandana'),
    (r'shoonya',                                                     'shoonya intensive'),
    (r'samyama',                                                     'samyama'),
    (r'vairagya',                                                    'vairagya'),
    (r'eye care.*(upayoga|upa yoga)',                                 'eye care upa yoga'),
    (r'eye care.*(shanmukhi)',                                        'eye care shanmukhi'),
    (r'eye care',                                                    'eye care'),
    (r'shanmukhi',                                                   'shanmukhi'),
    (r'angamardana',                                                 'angamardana'),
    (r'surya kriya.*surya shakti|surya shakti.*surya kriya',          'surya kriya surya shakti'),
    (r'surya kriya',                                                 'surya kriya'),
    (r'surya shakti',                                                'surya shakti'),
    (r'yogasanas|yoga asana',                                        'yogasanas'),
    (r'bhuta shuddhi',                                               'bhuta shuddhi'),
    (r'bhastrika',                                                   'bhastrika'),
    (r'hatha yoga for children|hatha children',                      'hatha children'),
    (r'hatha yoga',                                                  'hatha yoga'),
    (r'thoppukarnam',                                                'thoppukarnam'),
    (r'jala neti',                                                   'jala neti'),
    (r'sutra neti',                                                  'sutra neti'),
    (r'nauli',                                                       'nauli'),
    (r'kapalbhati',                                                  'kapalbhati'),
    (r'trataka',                                                     'trataka'),
    (r'guru purnima',                                                'guru purnima'),
    (r'volunteers meet|volunteer meet',                              'volunteers meet'),
    (r'guru pooja|guru puja',                                        'guru pooja'),
    (r'upa yoga|upayoga',                                            'upa yoga'),
    (r'isha kriya',                                                  'isha kriya'),
    (r'chit shakti',                                                 'chit shakti'),
    (r'nada aradhana',                                               'nada aradhana'),
    (r'satsang|sathsang',                                            'satsang'),
    (r'isha janani',                                                 'isha janani'),
]

# Maps python centre key → CENTRE_DATA / MONTHLY_DATA key in HTML
CENTRE_KEY_MAP = {
    'electronic':    'Electronic City',
    'banaswadi':     'Banaswadi',
    'hebbal':        'Hebbal',
    'indiranagar':   'Indiranagar',
    'jayanagar':     'IP - Jayanagar',
    'malleswaram':   'IP - Malleswaram',
    'marathahalli':  'IP - Marathahalli',
    'vijayanagar':   'IP - Vijayanagar',
    'budigere cross':'Budigere Cross',
    'mangalore':     'Mangalore',
    'whitefield':    'Whitefield',
    'mysuru':        'Mysuru',
    'bg road':       'Bannerghatta Road',
    'ballari':       'Ballari',
    'belagavi':      'Belagavi',
    'hubbali':       'Hubbali',
    'koramangala':   'Koramangala',
    'kanakapura':    'Kanakapura Road',
    'girinagar':     'Girinagar',
    'hassan':        'Hassan',
    'chikkaballapura': 'Chikkaballapura',
    'chikballapur':    'Chikkaballapura',
    'mandya':          'Mandya',
    'shivamogga':      'Shivamogga',
    'shimoga':         'Shivamogga',
    'tumkur':          'Tumkur',
    'udupi':           'Udupi',
    'kodagu':          'Kodagu',
}

# Maps normalised prog key → CENTRE_DATA / MONTHLY_DATA sub-key
PROG_TO_CD_KEY = {
    'inner engineering': 'ie',
    'bhava spandana':    'bsp',
    'shoonya intensive': 'shoonya',
    'samyama':           'samyama',
}

_MONTH_MAP = {
    'jan':'01','feb':'02','mar':'03','apr':'04','may':'05','jun':'06',
    'jul':'07','aug':'08','sep':'09','oct':'10','nov':'11','dec':'12'
}
_LANG_RE = re.compile(r'\b(English|Kannada|Tamil|Hindi|Telugu|Malayalam|Marathi)\b', re.IGNORECASE)

def extract_start_date(text):
    t = text.lower()
    # Pattern 0: cross-month range "29 Oct - 01 Nov 2026" or "29 Oct – 01 Nov 2026"
    # The year sits at the end with the end-month; we extract the start day+month.
    m = re.search(
        r'\b(\d{1,2})\s+(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s*[-–]\s*\d{1,2}\s+\w+\s+(20\d{2})\b',
        t)
    if m:
        return f"{m.group(3)}-{_MONTH_MAP[m.group(2)]}-{m.group(1).zfill(2)}"
    # Pattern 1: "12-15 Nov 2026" or "12 Nov 2026" — day-first, same-month range.
    # Requires a digit before the month name so "marathahalli" can never match.
    m = re.search(
        r'\b(\d{1,2})(?:-\d{1,2})?\s+(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+(20\d{2})\b',
        t)
    if m:
        return f"{m.group(3)}-{_MONTH_MAP[m.group(2)]}-{m.group(1).zfill(2)}"
    # Pattern 2 (fallback): "Nov 12 2026" or "Nov 12, 2026" — month-first format.
    m  = re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+(\d{1,2})\b', t)
    ym = re.search(r'\b(20\d{2})\b', text)
    if m and ym:
        return f"{ym.group(1)}-{_MONTH_MAP[m.group(1)]}-{m.group(2).zfill(2)}"
    return None

def norm_centre(text):
    t = text.lower()
    for pat, key in CENTRE_PATTERNS:
        if re.search(pat, t): return key
    return None

def norm_prog(text):
    t = text.lower()
    for pat, key in PROG_PATTERNS:
        if re.search(pat, t): return key
    return None

# ── Find all eligible files recursively in Santhosha Data ────────────────────
def find_all_files():
    candidates = (
        glob.glob(os.path.join(SANTHOSHA_DIR, "**", "*.xlsx"),    recursive=True) +
        glob.glob(os.path.join(SANTHOSHA_DIR, "**", "*.numbers"), recursive=True) +
        glob.glob(os.path.join(SANTHOSHA_DIR, "**", "*.xls"),     recursive=True)
    )
    candidates = [f for f in candidates
                  if not os.path.basename(f).startswith('~$')
                  and not os.path.basename(f).startswith('.')]
    if not candidates:
        if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
            return [sys.argv[1]]
        print(f"✗ No Numbers/Excel files found in: {SANTHOSHA_DIR}")
        sys.exit(1)
    candidates.sort(key=os.path.getmtime, reverse=True)
    return candidates

# ── Export .numbers → CSV via AppleScript ────────────────────────────────────
def export_numbers_to_csv(numbers_path, tmp_csv):
    script = f'''
tell application "Numbers"
    open POSIX file "{numbers_path}"
    delay 2
    set theDoc to front document
    export theDoc to POSIX file "{tmp_csv}" as CSV
    close theDoc saving no
end tell
'''
    r = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    if r.returncode != 0:
        print(f"  ✗ AppleScript error: {r.stderr.strip()}")
        return False
    if not os.path.exists(tmp_csv):
        print(f"  ✗ CSV not created — is Numbers installed?")
        return False
    return True

# ── Detect and parse CRM "Total row" xlsx format ─────────────────────────────
# Format: Row1=labels, Row2=years, Row3="Count" headers, Row4="Total" with annual counts
# Row5+: detail rows — either "Month Year" (EC-style) or "Year" (other centres)
def try_parse_crm_xlsx(xlsx_path, regs, cd_updates, monthly_updates, source_label, file_timestamps=None):
    """Returns True if file matches CRM format and was parsed; False to fall back."""
    if not HAS_OPENPYXL:
        return False
    try:
        wb   = openpyxl.load_workbook(xlsx_path, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
    except Exception as e:
        print(f"  ✗ openpyxl error: {e}")
        return False

    # Detect CRM format: row[1] has 4-digit year strings, row[3][0] == 'Total'
    if len(rows) < 4:
        return False
    years_row = rows[1]
    total_row = rows[3]
    if str(total_row[0]).strip().lower() != 'total':
        return False
    year_cols = [(i, str(v).strip()) for i, v in enumerate(years_row)
                 if v and re.match(r'^20\d{2}$|^1\d{3}$', str(v).strip())]
    if not year_cols:
        return False

    # Infer centre from parent folder name
    parent_folder = os.path.basename(os.path.dirname(xlsx_path))
    centre_key    = norm_centre(parent_folder)
    html_centre   = CENTRE_KEY_MAP.get(centre_key) if centre_key else None

    # Infer programme from filename (normalise underscores → spaces for pattern matching)
    fname    = os.path.basename(xlsx_path).lower().replace('_', ' ')
    prog_key = norm_prog(fname)
    cd_key   = PROG_TO_CD_KEY.get(prog_key) if prog_key else None

    if not html_centre or not cd_key:
        print(f"  ⚠ CRM format detected but couldn't infer centre ({parent_folder!r}) "
              f"or programme ({fname!r}) — skipping")
        return True  # still CRM format, just not updatable

    # Record per-centre file timestamp (keep the latest mtime if multiple files per centre)
    if file_timestamps is not None:
        mtime = os.path.getmtime(xlsx_path)
        if html_centre not in file_timestamps or mtime > file_timestamps[html_centre]:
            file_timestamps[html_centre] = mtime

    # Build col_index → year map for quick lookup
    col_to_year = {i: yr for i, yr in year_cols}

    # ── Annual totals → CENTRE_DATA ──────────────────────────────────────────
    # For IE: the Total row is programme-specific, use it directly.
    # For BSP/Shoonya/Samyama: the Total row is an all-member aggregate (same
    # value across all programme files), so we skip it and instead derive annual
    # totals by summing the programme-specific monthly rows (see section below).
    matched_annual = 0
    if cd_key == 'ie':
        for col_i, year in year_cols:
            val = total_row[col_i]
            if val is None:
                continue
            try:
                count = int(float(str(val)))
            except:
                continue
            if count <= 0:
                continue
            cd_updates.setdefault(html_centre, {}).setdefault(cd_key, {})[year] = count
            matched_annual += 1

    # ── Monthly breakdown → MONTHLY_DATA (only if row labels contain month names) ──
    # EC-style: "     January 2026", "     March 2026" etc.
    # Other:    "     2026" — year only, no monthly data available
    matched_monthly = 0
    month_name_re = re.compile(
        r'\b(january|february|march|april|may|june|july|august|september|october|november|december)\b',
        re.IGNORECASE
    )
    month_num = {
        'january':'01','february':'02','march':'03','april':'04',
        'may':'05','june':'06','july':'07','august':'08',
        'september':'09','october':'10','november':'11','december':'12'
    }

    for row in rows[4:]:
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        mn = month_name_re.search(label)
        if not mn:
            continue  # year-only row — no monthly breakdown available
        month_str = month_num[mn.group(1).lower()]
        # Extract year from label (e.g. "January 2026" → 2026)
        yr_m = re.search(r'\b(20\d{2})\b', label)
        if not yr_m:
            continue
        year = yr_m.group(1)

        # Sum ALL year columns for this row — the CRM pivot columns are member-join-year,
        # but the row label year is when the programme happened. Off-diagonal entries
        # (join year ≠ programme year) are the majority for repeat programmes like BSP/Shoonya.
        # Summing all columns gives the true participant count for that month.
        count = 0
        for col_i, col_yr in year_cols:
            val = row[col_i]
            if val is None:
                continue
            try:
                count += int(float(str(val)))
            except:
                pass
        if count <= 0:
            continue

        (monthly_updates
            .setdefault(html_centre, {})
            .setdefault(cd_key, {})
            .setdefault(year, {})[month_str]) = count
        matched_monthly += 1

    # ── Derive annual totals from monthly rows for BSP/Shoonya/Samyama ──────
    if cd_key != 'ie' and matched_monthly > 0:
        centre_monthly = monthly_updates.get(html_centre, {}).get(cd_key, {})
        for yr, months in centre_monthly.items():
            yr_total = sum(v for v in months.values() if v)
            if yr_total > 0:
                cd_updates.setdefault(html_centre, {}).setdefault(cd_key, {})[yr] = yr_total
                matched_annual += 1

    monthly_note = f", {matched_monthly} monthly entries" if matched_monthly else ""
    print(f"  ✓ {source_label} [CRM]: {matched_annual} annual{monthly_note} → "
          f"{html_centre} / {cd_key}")
    return True

# ── Read .xlsx via openpyxl → temp CSV (fallback for non-CRM xlsx) ───────────
def export_xlsx_to_csv(xlsx_path, tmp_csv):
    if not HAS_OPENPYXL:
        print(f"  ⚠ openpyxl not installed — falling back to AppleScript")
        return export_numbers_to_csv(xlsx_path, tmp_csv)
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        with open(tmp_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(['' if v is None else str(v) for v in row])
        return True
    except Exception as e:
        print(f"  ✗ openpyxl error: {e}")
        return False

# ── Route to correct exporter based on file type ─────────────────────────────
def export_to_csv(file_path, tmp_csv):
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        return export_xlsx_to_csv(file_path, tmp_csv)
    else:
        return export_numbers_to_csv(file_path, tmp_csv)

# ── Parse one CSV, merge into regs (LIVE_REGS) and cd_updates (CENTRE_DATA) ──
def parse_csv(csv_path, regs, cd_updates, source_label, lang_map=None):
    skipped = 0
    matched = 0

    with open(csv_path, newline='', encoding='utf-8-sig', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            if not row: continue
            name = row[0].strip()
            count_str = ''
            for col in row[1:]:
                col = col.strip()
                if col and re.match(r'^-?\d+(\.\d+)?$', col):
                    count_str = col
                    break
            if not name or not count_str:
                continue
            try:
                count = int(float(count_str))
            except:
                continue
            if count < 0:
                continue

            centre = norm_centre(name)
            prog   = norm_prog(name)
            if not centre or not prog:
                skipped += 1
                continue

            # ── LIVE_REGS update ────────────────────────────────────────────
            date     = extract_start_date(name)
            prog_key = f"{prog}|{date}" if date else prog
            if centre not in regs:
                regs[centre] = {}
            if lang_map is not None:
                _lm = _LANG_RE.search(name)
                if _lm and date:
                    lang_map.setdefault(centre, {})[f"{prog}|{date}"] = _lm.group(1).title()
            if date:
                regs[centre][prog_key] = count
            else:
                regs[centre][prog_key] = regs[centre].get(prog_key, 0) + count

            # ── CENTRE_DATA update (IE / BSP / Shoonya / Samyama only) ─────
            cd_key      = PROG_TO_CD_KEY.get(prog)
            html_centre = CENTRE_KEY_MAP.get(centre)
            if cd_key and html_centre:
                year = (date[:4] if date else None) or CURRENT_YEAR
                if html_centre not in cd_updates:
                    cd_updates[html_centre] = {}
                if cd_key not in cd_updates[html_centre]:
                    cd_updates[html_centre][cd_key] = {}
                cd_updates[html_centre][cd_key][year] = (
                    cd_updates[html_centre][cd_key].get(year, 0) + count
                )

            matched += 1

    print(f"  ✓ {source_label}: {matched} matched, {skipped} skipped")
    return matched

# ── Read existing CENTRE_DATA from HTML ──────────────────────────────────────
CD_START = "/* __CENTRE_DATA_START__ */"
CD_END   = "/* __CENTRE_DATA_END__ */"

def _js_to_json(s):
    """Convert JS object literal (single-quoted keys, bare keys) to valid JSON."""
    # Quote bare JS keys (word chars before colon, not already quoted)
    s = re.sub(r"(?<!['\"\w])(\b[a-zA-Z_]\w*)\s*:", r'"\1":', s)
    # Replace single-quoted strings with double-quoted
    s = re.sub(r"'([^']*)'", r'"\1"', s)
    return s

def read_centre_data(html):
    m = re.search(re.escape(CD_START) + r'(.*?)' + re.escape(CD_END), html, re.DOTALL)
    if not m:
        return None, None
    block = m.group(1).strip()
    obj_m = re.search(r'const CENTRE_DATA\s*=\s*(\{.*\});', block, re.DOTALL)
    if not obj_m:
        return None, None
    try:
        return json.loads(obj_m.group(1)), block
    except json.JSONDecodeError:
        try:
            return json.loads(_js_to_json(obj_m.group(1))), block
        except json.JSONDecodeError as e:
            print(f"  ✗ Could not parse CENTRE_DATA: {e}")
            return None, None

# ── Read existing MONTHLY_DATA from HTML ─────────────────────────────────────
MD_START = "/* __MONTHLY_DATA_START__ */"
MD_END   = "/* __MONTHLY_DATA_END__ */"

def read_monthly_data(html):
    m = re.search(re.escape(MD_START) + r'(.*?)' + re.escape(MD_END), html, re.DOTALL)
    if not m:
        return None
    block = m.group(1)
    js = re.search(r'const MONTHLY_DATA\s*=\s*(\{.*\});', block, re.DOTALL)
    if not js:
        return None
    try:
        return json.loads(js.group(1))
    except json.JSONDecodeError:
        try:
            return json.loads(_js_to_json(js.group(1)))
        except json.JSONDecodeError as e:
            print(f"  ✗ Could not parse MONTHLY_DATA: {e}")
            return None

# ── Build JS object string for MONTHLY_DATA (compact inner dicts) ────────────
def monthly_data_to_js(md):
    """Serialize MONTHLY_DATA back to compact JS object literal."""
    lines = ['const MONTHLY_DATA = {']
    centres = sorted(md.keys())
    for ci, centre in enumerate(centres):
        comma_c = ',' if ci < len(centres) - 1 else ''
        lines.append(f"  '{centre}': {{")
        progs = md[centre]
        prog_keys = list(progs.keys())
        for pi, prog in enumerate(prog_keys):
            comma_p = ',' if pi < len(prog_keys) - 1 else ''
            lines.append(f"    {prog}: {{")
            years = sorted(progs[prog].keys())
            for yi, year in enumerate(years):
                months = progs[prog][year]
                comma_y = ',' if yi < len(years) - 1 else ''
                inner = ','.join(f"'{m}':{v}" for m, v in sorted(months.items()))
                lines.append(f"      '{year}':{{{inner}}}{comma_y}")
            lines.append(f"    }}{comma_p}")
        lines.append(f"  }}{comma_c}")
    lines.append('};')
    return '\n'.join(lines)

# ── IE Online data parsing ────────────────────────────────────────────────────
IEO_START = "/* __IE_ONLINE_START__ */"
IEO_END   = "/* __IE_ONLINE_END__ */"

ALL_IEO_STATUSES = [
    'Not Registered','Not Started','Started',
    'Step 1 Completed','Step 2 Completed','Step 3 Completed',
    'Step 4 Completed','Step 5 Completed','Step 6 Completed',
    'Course Completed',
]
IEO_SKIP_CENTRES = {'None','Karnataka - Others','Vijayanagara'}

def parse_ie_online(xlsx_path):
    """Parse 'IE Online - All center.xlsx' pivot table.
    Returns dict: centre → {years: {year → {s, cc, st, ns, months: {month → {s, cc, st, ns}}}}}
    """
    try:
        import openpyxl as _opx
        wb = _opx.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"  ⚠ Could not open IE Online file: {e}")
        return {}

    # raw[centre][year][month][status] = count
    # xlsx hierarchy: Year(indent5) → Month(indent10) → Centre(indent15) → Status(indent20)
    raw = {}
    cur_year = cur_month = cur_centre = None

    for row in ws.iter_rows(values_only=True):
        v0 = row[0]
        if v0 is None: continue
        s_raw = str(v0); s = s_raw.strip()
        if not s or s == 'Total': continue
        indent = len(s_raw) - len(s)
        # Grand total is always the last column in the row
        try: count = int(row[-1] or 0)
        except: count = 0

        if indent == 5:     # Year  e.g. "2023"
            cur_year = s; cur_month = None; cur_centre = None
        elif indent == 10:  # Month e.g. "October 2023"
            cur_month = s; cur_centre = None
        elif indent == 15:  # Centre e.g. "Hebbal"
            cur_centre = s
            raw.setdefault(s, {}).setdefault(cur_year, {}).setdefault(cur_month, {st: 0 for st in ALL_IEO_STATUSES})
        elif indent == 20:  # Status e.g. "Course Completed"
            if cur_centre and cur_year and cur_month:
                m_dict = raw.get(cur_centre, {}).get(cur_year, {}).get(cur_month, {})
                if s in m_dict:
                    m_dict[s] += count

    def _to_entry(yd):
        steps = [yd.get(f'Step {i} Completed', 0) for i in range(1, 7)]
        return {'s': steps, 'cc': yd.get('Course Completed', 0),
                'st': yd.get('Started', 0),
                'ns': yd.get('Not Started', 0) + yd.get('Not Registered', 0)}

    out = {}
    for centre, yrs in raw.items():
        if centre in IEO_SKIP_CENTRES: continue
        out[centre] = {'years': {}}
        for yr, months in yrs.items():
            # aggregate year totals from months
            yr_agg = {st: 0 for st in ALL_IEO_STATUSES}
            months_out = {}
            for mon, md in months.items():
                for st, v in md.items(): yr_agg[st] += v
                months_out[mon] = _to_entry(md)
            entry = _to_entry(yr_agg)
            entry['months'] = months_out
            out[centre]['years'][yr] = entry

    print(f"  ✓ IE Online: {len(out)} centres parsed")
    return out


# ── IE Online current month parsing ───────────────────────────────────────────
IEO_CM_START = "/* __IEO_CURRENT_MONTH_START__ */"
IEO_CM_END   = "/* __IEO_CURRENT_MONTH_END__ */"

def parse_ieo_current_month(xlsx_path):
    """Parse IE_Online_CurrentMonthStatus_*.xlsx.
    Returns dict: {month: 'August 2026', centres: {centre: count}, total: N}
    """
    import re as _re
    try:
        import openpyxl as _opx
        wb = _opx.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"  ⚠ Could not open current month IE Online file: {e}")
        return {}

    # Extract month from filename e.g. IE_Online_CurrentMonthStatus_August.xlsx
    fname = os.path.basename(xlsx_path)
    m = _re.search(r'CurrentMonthStatus[_\-]([A-Za-z]+)', fname)
    month_name = m.group(1) if m else 'Current'
    import datetime as _dt
    month_str = f"{month_name} {_dt.date.today().year}"

    centres = {}
    total = 0
    for row in ws.iter_rows(values_only=True):
        v0, v1 = row[0], row[1]
        if v0 is None or v1 is None: continue
        s_raw = str(v0)
        s = s_raw.strip()
        if not s or s in ('Total', 'Karnataka', 'None', 'Count'): continue
        try: count = int(v1)
        except: continue
        indent = len(s_raw) - len(s)
        if indent >= 10:   # individual centre rows
            centres[s] = count
            total += count

    print(f"  ✓ IE Online current month ({month_str}): {len(centres)} centres, {total} total")
    return {'month': month_str, 'centres': centres, 'total': total}


# ── Fetch upcoming programs from Isha API and build embedded data ─────────────
UP_START = "/* __UPCOMING_DATA_START__ */"
UP_END   = "/* __UPCOMING_DATA_END__ */"

PLACE_CENTRE = {
    'vijayanagar':     'IP - Vijayanagar',
    'jayanagar':       'IP - Jayanagar',
    'jp nagar':        'IP - Jayanagar',
    'marathahalli':    'IP - Marathahalli',
    'marathali':       'IP - Marathahalli',
    'malleswaram':     'IP - Malleswaram',
    'malleswar':       'IP - Malleswaram',
    'malleshwaram':    'IP - Malleswaram',
    'electronic city': 'Electronic City',
    'ecity':           'Electronic City',
    'e city':          'Electronic City',
    'banaswadi':       'Banaswadi',
    'banasawadi':      'Banaswadi',
    'hebbal':          'Hebbal',
    'hebbala':         'Hebbal',
    'indiranagar':     'Indiranagar',
    'indira nagar':    'Indiranagar',
    'sarjapur':        'Sarjapur Road',
    'sargapur':        'Sarjapur Road',
    'kanakapura':      'Kanakapura Road',
    'kanakapur':       'Kanakapura Road',
    'hsr layout':      'HSR Layout',
    'hsr':             'HSR Layout',
    'mysuru':          'Mysuru',
    'mysore':          'Mysuru',
    'sadhguru sannidhi': 'Sadhguru Sannidhi',
    'sannidhi':          'Sadhguru Sannidhi',
    'koramangala':     'Koramangala',
    'bannerghatta road': 'Bannerghatta Road',
    'bannerghatta':    'Bannerghatta Road',
    'bg road':         'Bannerghatta Road',
    'yelahanka':       'Yelahanka',
    'girinagar':       'Girinagar',
    'hassan':          'Hassan',
    'kengeri':         'Kengeri',
    'peenya':          'Peenya',
    'chandapura':      'Chandapura',
    'begur':           'Begur',
    'budigere cross':  'Budigere Cross',
    'budigere':        'Budigere Cross',
    'mangalore':       'Mangalore',
    'mangaluru':       'Mangalore',
    'hubballi':        'Hubbali',
    'hubli':           'Hubbali',
    'ballari':         'Ballari',
    'bellary':         'Ballari',
    'belagavi':        'Belagavi',
    'belgaum':         'Belagavi',
    'dharwad':         'Dharwad',
    'chikkaballapura': 'Chikkaballapura',
    'chikballapur':    'Chikkaballapura',
    'mandya':          'Mandya',
    'shivamogga':      'Shivamogga',
    'shimoga':         'Shivamogga',
    'tumkur':          'Tumkur',
    'tumakuru':        'Tumkur',
    'udupi':           'Udupi',
    'whitefield':      'Whitefield',
    'singasandra':     'Singasandra',
}

KA_CITIES = [
    'bengaluru','bangalore','mysuru','mysore','hubli','hubbali','hubballi','dharwad',
    'mangaluru','mangalore','belagavi','belgaum','kalaburagi','gulbarga','tumkur','tumakuru',
    'davanagere','bellary','ballari','shivamogga','shimoga','udupi','hassan','mandya','raichur',
    'bidar','vijayapura','bijapur','chitradurga','chikkamagaluru','chikmagalur','koppal','gadag',
    'bagalkot','yadgir','chamarajanagar','kodagu','coorg','kolar','chikkaballapur','ramanagara',
    'kanakapura','karnataka',
]
IYC_IDENTIFIERS = ['velliangiri','isha yoga center','isha yoga centre']
TN_PROG_KEYWORDS = ['bhava spandana','shoonya','samyama','vairagya','upa yoga','surya shakti']

IE_IMG      = 'https://static.sadhguru.org/d/46272/1695654981-ieo2023_sg-banner_website.jpg'
BSP_IMG     = 'https://static.sadhguru.org/d/46272/1650519638-website-thumbnail-yogameditation-bsp.jpg'
SAMYAMA_IMG = 'https://static.sadhguru.org/d/46272/1650517087-website-thumbnail-yogameditation-samyama.jpg'
SHOONYA_IMG = 'https://static.sadhguru.org/d/46272/1650449300-website-thumbnail-yogameditation-shoonya.jpg'

def _fmt_date_range(fr, to):
    """Replicate JS _fmtDateRange for Python."""
    MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    try:
        a = datetime.datetime.fromisoformat((fr or '').replace(' ', 'T').split('T')[0])
        b = datetime.datetime.fromisoformat(((to or fr) or '').replace(' ', 'T').split('T')[0])
    except Exception:
        return fr or ''
    if a.month == b.month and a.year == b.year:
        return f"{a.day}–{b.day} {MONTHS[a.month-1]} {a.year}"
    if a.year == b.year:
        return f"{a.day} {MONTHS[a.month-1]}–{b.day} {MONTHS[b.month-1]} {a.year}"
    return f"{a.day} {MONTHS[a.month-1]} {a.year}–{b.day} {MONTHS[b.month-1]} {b.year}"

def _ie_name(n):
    l = n.lower()
    if not (l.find('inner engineering') >= 0 or l.find('isha yoga') >= 0 or
            l.find('ie total') >= 0 or l.find('inner eng') >= 0):
        return None
    if 'retreat'  in l: return 'Inner Engineering Retreat'
    if '4 day'    in l or '4day' in l: return 'Inner Engineering 4 Days'
    return 'Inner Engineering 7 Days'

def _hatha_name(n):
    l = n.lower()
    if 'eye care' in l and ('upa yoga' in l or 'upayoga' in l): return 'Eye Care & Upa Yoga'
    if 'eye care' in l and ('shanmukhi' in l or 'mudra' in l):  return 'Eye Care Yoga & Shanmukhi Mudra'
    if 'eye care'     in l: return 'Eye Care'
    if 'surya kriya'  in l and 'surya shakti' in l: return 'Surya Kriya & Surya Shakti'
    if 'surya kriya'  in l and 'weekend' in l:      return 'Surya Kriya Weekend'
    if 'surya kriya'  in l: return 'Surya Kriya'
    if 'angamardana'  in l: return 'Angamardana'
    if 'yogasanas'    in l or 'yoga asana' in l: return 'Yogasanas'
    if 'bhuta shud'   in l: return 'Bhuta Shuddhi'
    if 'shanmukhi'    in l: return 'Shanmukhi'
    if 'thoppukarnam' in l: return 'Thoppukarnam'
    if 'jala neti'    in l: return 'Jala Neti'
    if 'sutra neti'   in l: return 'Sutra Neti'
    if 'nauli'        in l: return 'Nauli'
    if 'kapalbhati'   in l: return 'Kapalbhati'
    if 'trataka'      in l: return 'Trataka'
    return None

def _match_centre(p):
    combined = ((p.get('place') or '') + ' ' + (p.get('city') or '')).lower()
    # Must check longer keys first to avoid 'jayanagar' matching inside 'vijayanagar'
    for k in sorted(PLACE_CENTRE, key=len, reverse=True):
        if k in combined:
            return PLACE_CENTRE[k]
    name_l = (p.get('name') or '').lower()
    for k in sorted(PLACE_CENTRE, key=len, reverse=True):
        if k in name_l:
            return PLACE_CENTRE[k]
    return None

def _is_included_region(p):
    pin = str(p.get('pin') or p.get('pincode') or p.get('zip') or '').strip()
    st  = (p.get('state') or p.get('st') or p.get('region') or '').lower()
    cp  = ((p.get('city') or '') + (p.get('place') or '')).lower()
    nl  = (p.get('name') or '').lower()
    if pin and re.match(r'^5[6-9]', pin): return True
    if st and ('karnataka' in st or st == 'ka' or st == 'kk'): return True
    if 'karnataka' in cp or any(k in cp for k in KA_CITIES): return True
    is_iyc = any(k in cp for k in IYC_IDENTIFIERS)
    if is_iyc and any(k in nl for k in TN_PROG_KEYWORDS): return True
    return False

def detail_url(id_):
    return f"https://isha.sadhguru.org/in/en/program-details?id={id_}"

def fetch_upcoming_programs():
    """Fetch from Isha API and return (ie_up, bsp_up, shoonya_up, samyama_up, hatha_up, all_up).
    Returns None on failure."""
    API_URL = 'https://api.ishafoundation.org/scheduleApi/data.php?task=list&activeFilter=100'
    try:
        req = urllib.request.Request(API_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=20) as r:
            data = json.loads(r.read().decode('utf-8'))
    except Exception as e:
        print(f"  ⚠ Could not fetch Isha API: {e}")
        return None

    results = data.get('results', {})
    if not isinstance(results, dict):
        print("  ⚠ Unexpected API response shape")
        return None

    today = datetime.date.today()
    all_progs = []
    for ok in results:
        inner = results[ok]
        if isinstance(inner, dict):
            for id_ in inner:
                all_progs.append(inner[id_])

    # Filter: exclude US, keep future (or dateless) programs
    future = []
    for p in all_progs:
        if p.get('cntry') == 'US':
            continue
        fr = p.get('fr')
        if not fr:
            future.append(p)
            continue
        try:
            start = datetime.date.fromisoformat(fr.split(' ')[0].split('T')[0])
            if start >= today:
                future.append(p)
        except Exception:
            future.append(p)

    # IE UPCOMING
    ie_up = {}
    for p in future:
        centre = _match_centre(p)
        if not centre: continue
        name = _ie_name(p.get('name',''))
        if not name: continue
        if centre not in ie_up: ie_up[centre] = []
        ie_up[centre].append({'name': name, 'date': _fmt_date_range(p.get('fr'), p.get('to')),
                              'url': detail_url(p.get('id','')), 'img': IE_IMG, 'fr': p.get('fr','')})

    # BSP / SHOONYA / SAMYAMA
    bsp_up     = []
    shoonya_up = []
    samyama_up = []
    for p in future:
        if not _is_included_region(p): continue
        nl = (p.get('name') or '').lower()
        entry = {'name': p.get('name',''), 'date': _fmt_date_range(p.get('fr'), p.get('to')),
                 'loc': ((p.get('place') or '') + ', ' if p.get('place') else '') + (p.get('city') or ''),
                 'url': detail_url(p.get('id','')), 'fr': p.get('fr','')}
        if 'bhava spandana' in nl: bsp_up.append(entry)
        if 'shoonya'        in nl: shoonya_up.append(entry)
        if 'samyama'        in nl: samyama_up.append(entry)

    # HATHA UPCOMING
    hatha_up = {}
    for p in future:
        centre = _match_centre(p)
        if not centre: continue
        name = _hatha_name(p.get('name',''))
        if not name: continue
        if centre not in hatha_up: hatha_up[centre] = []
        hatha_up[centre].append({'name': name, 'date': _fmt_date_range(p.get('fr'), p.get('to')),
                                 'url': detail_url(p.get('id','')), 'fr': p.get('fr','')})

    # ALL UPCOMING (Karnataka + TN/IYC)
    KNOWN_ORDER = ['Banaswadi','Electronic City','Hebbal','Indiranagar',
                   'IP - Jayanagar','IP - Malleswaram','IP - Marathahalli','IP - Vijayanagar',
                   'Kanakapura Road','Sarjapur Road']
    all_up_raw = {}
    for p in future:
        if not _is_included_region(p): continue
        matched = _match_centre(p)
        centre_key = matched or ', '.join(filter(None,[p.get('place'),p.get('city')])) or 'Karnataka'
        if centre_key not in all_up_raw: all_up_raw[centre_key] = []
        img_path = p.get('img','')
        img_url  = f"https://static.sadhguru.org/d{img_path}" if img_path else ''
        all_up_raw[centre_key].append({
            'name':  p.get('name',''),
            'date':  _fmt_date_range(p.get('fr'), p.get('to')),
            'loc':   ((p.get('place') or '') + ', ' if p.get('place') else '') + (p.get('city') or ''),
            'url':   detail_url(p.get('id','')),
            'fr':    p.get('fr',''),
            'to':    p.get('to',''),
            'img':   img_url,
            'place': p.get('place',''),
            'city':  p.get('city',''),
            'state': p.get('state') or p.get('st') or p.get('region') or '',
            'pin':   str(p.get('pin') or p.get('pincode') or p.get('zip') or '').strip(),
        })
    # ── SATSANG fetch (separate endpoint: api.php?category=173) ─────────────
    # The main data.php API does not return Karnataka satsang events.
    SAT_URL = ('https://api.ishafoundation.org/scheduleApi/api.php'
               '?option=com_program&v=2&format=json&task=filter'
               '&count=200&startrec=0&category=173')
    MONTH_MAP = {'jan':'01','feb':'02','mar':'03','apr':'04','may':'05','jun':'06',
                 'jul':'07','aug':'08','sep':'09','oct':'10','nov':'11','dec':'12'}
    KA_WORDS  = ['bengaluru','bangalore','karnataka','mangalore','mangaluru',
                 'mysuru','mysore','hubli','hubbali','hubballi','tumkur','tumakuru',
                 'kanakapura','belagavi','belgaum','ballari','bellary','dharwad',
                 'shivamogga','shimoga','udupi','hassan','mandya']

    def _parse_sat_date(s):
        """'5 Jul 2026' → '2026-07-05 00:00:00'"""
        m = re.search(r'(\d+)\s+([A-Za-z]+)\s+(\d{4})', s or '')
        if not m: return None
        mon = MONTH_MAP.get(m.group(2).lower()[:3])
        return f"{m.group(3)}-{mon}-{m.group(1).zfill(2)} 00:00:00" if mon else None

    sat_count = 0
    try:
        req2 = urllib.request.Request(SAT_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req2, timeout=20) as r2:
            sat_data = json.loads(r2.read().decode('utf-8'))
        # results is a list of {id: satsang_obj} dicts
        all_sat = []
        sat_results = sat_data.get('results', [])
        if isinstance(sat_results, list):
            for obj in sat_results:
                if isinstance(obj, dict):
                    all_sat.extend(obj.values())
        elif isinstance(sat_results, dict):
            for obj in sat_results.values():
                if isinstance(obj, dict):
                    all_sat.extend(obj.values())

        for s in all_sat:
            addr_l = (s.get('address') or '').lower()
            if not any(w in addr_l for w in KA_WORDS): continue
            fr = _parse_sat_date(s.get('date') or '')
            if fr:
                try:
                    if datetime.date.fromisoformat(fr.split(' ')[0]) < today: continue
                except Exception:
                    pass
            # Match to a specific centre
            combined = ((s.get('title') or '') + ' ' + addr_l).lower()
            centre_key = None
            for k in sorted(PLACE_CENTRE, key=len, reverse=True):
                if k in combined:
                    centre_key = PLACE_CENTRE[k]
                    break
            if not centre_key: continue
            if centre_key not in all_up_raw: all_up_raw[centre_key] = []
            # Skip if a satsang entry already exists for this centre
            if any('satsang' in (e.get('name') or '').lower() for e in all_up_raw[centre_key]):
                continue
            all_up_raw[centre_key].append({
                'name':  s.get('title') or 'Monthly Satsang',
                'date':  s.get('date') or '',
                'loc':   s.get('address') or '',
                'url':   (s.get('register_url') or
                          f"https://isha.sadhguru.org/in/en/program-details?id={s.get('program_id','')}"),
                'fr':    fr or '',
                'to':    fr or '',
                'img':   '',
                'place': s.get('address_title') or '',
                'city':  '',
                'state': 'Karnataka',
                'pin':   '',
            })
            sat_count += 1
        print(f"  ✓ Fetched satsang: {sat_count} Karnataka entries")
    except Exception as e:
        print(f"  ⚠ Could not fetch satsang programs: {e}")

    sorted_keys = sorted(all_up_raw.keys(),
        key=lambda k: (KNOWN_ORDER.index(k) if k in KNOWN_ORDER else len(KNOWN_ORDER), k))
    all_up = {k: all_up_raw[k] for k in sorted_keys}

    total = sum(len(v) for v in ie_up.values())
    print(f"  ✓ Fetched upcoming programs: {total} IE, {len(bsp_up)} BSP, "
          f"{len(shoonya_up)} Shoonya, {len(samyama_up)} Samyama, "
          f"{sum(len(v) for v in hatha_up.values())} Hatha, {sat_count} Satsang")
    return ie_up, bsp_up, shoonya_up, samyama_up, hatha_up, all_up

# ── Inject LIVE_REGS, CENTRE_DATA, MONTHLY_DATA, and CENTRE_TIMESTAMPS into HTML ─
LIVE_START = "/* __LIVE_REGS_START__ */"
LIVE_END   = "/* __LIVE_REGS_END__ */"
RH_START   = "/* __REGS_HISTORY_START__ */"
RH_END     = "/* __REGS_HISTORY_END__ */"
TS_START   = "/* __CENTRE_TIMESTAMPS_START__ */"
TS_END     = "/* __CENTRE_TIMESTAMPS_END__ */"

def read_regs_history(html):
    """Read existing LIVE_REGS_HISTORY array from HTML. Returns list of {date, regs} dicts."""
    m = re.search(re.escape(RH_START) + r'(.*?)' + re.escape(RH_END), html, re.DOTALL)
    if not m:
        return []
    block = m.group(1)
    arr_m = re.search(r'const LIVE_REGS_HISTORY\s*=\s*(\[.*?\]);', block, re.DOTALL)
    if not arr_m:
        return []
    try:
        return json.loads(arr_m.group(1))
    except json.JSONDecodeError:
        return []

def inject_html(html, regs, centre_data=None, monthly_data=None, latest_mtime=None, file_timestamps=None, upcoming=None, ieo_data=None, ieo_cm_data=None, regs_history=None, lang_map=None):
    # 1. LIVE_REGS block
    dt          = datetime.datetime.fromtimestamp(latest_mtime)
    updated_str = dt.strftime('%d %b %Y, %I:%M %p')
    regs_json   = json.dumps(regs, indent=2, ensure_ascii=False)
    lang_json   = json.dumps(lang_map or {}, indent=2, ensure_ascii=False)
    live_block  = (f"{LIVE_START}\nconst LIVE_REGS = {regs_json};\nconst LIVE_REGS_LANG = {lang_json};\n"
                   f"const LIVE_REGS_UPDATED = {json.dumps(updated_str)};\n{LIVE_END}")

    if LIVE_START not in html:
        print("  ✗ LIVE_REGS markers not found in HTML")
        sys.exit(1)
    html = re.sub(re.escape(LIVE_START) + r'.*?' + re.escape(LIVE_END),
                  live_block, html, flags=re.DOTALL)
    print(f"  ✓ Updated LIVE_REGS  (as of {updated_str})")

    # 1b. REGS_HISTORY block — keep last 3 daily snapshots
    if regs_history is not None and RH_START in html:
        rh_block = (f"{RH_START}\nconst LIVE_REGS_HISTORY = "
                    f"{json.dumps(regs_history, ensure_ascii=False)};\n{RH_END}")
        html = re.sub(re.escape(RH_START) + r'.*?' + re.escape(RH_END),
                      rh_block, html, flags=re.DOTALL)
        print(f"  ✓ Updated LIVE_REGS_HISTORY ({len(regs_history)} day(s))")
    elif RH_START not in html:
        print(f"  ⚠ REGS_HISTORY markers not found — skipping")

    # 2. CENTRE_DATA block
    if centre_data is not None and CD_START in html:
        cd_json  = json.dumps(centre_data, indent=2, ensure_ascii=False)
        cd_block = f"{CD_START}\nconst CENTRE_DATA = {cd_json};\n{CD_END}"
        html = re.sub(re.escape(CD_START) + r'.*?' + re.escape(CD_END),
                      cd_block, html, flags=re.DOTALL)
        print(f"  ✓ Updated CENTRE_DATA")
    else:
        print(f"  ⚠ CENTRE_DATA markers not found — skipping")

    # 3. MONTHLY_DATA block
    if monthly_data is not None and MD_START in html:
        md_js    = monthly_data_to_js(monthly_data)
        md_block = f"{MD_START}\n// Monthly data — centres with month-level CRM data (others have annual totals only)\n{md_js}\n{MD_END}"
        html = re.sub(re.escape(MD_START) + r'.*?' + re.escape(MD_END),
                      md_block, html, flags=re.DOTALL)
        print(f"  ✓ Updated MONTHLY_DATA")
    else:
        print(f"  ⚠ MONTHLY_DATA markers not found — skipping")

    # 4. CENTRE_TIMESTAMPS block
    if file_timestamps is not None and TS_START in html:
        # Convert mtime floats → human-readable date strings
        ts_dict = {
            centre: datetime.datetime.fromtimestamp(mtime).strftime('%d %b %Y')
            for centre, mtime in file_timestamps.items()
        }
        ts_json  = json.dumps(ts_dict, indent=2, ensure_ascii=False)
        ts_block = f"{TS_START}\nconst CENTRE_DATA_TIMESTAMPS = {ts_json};\n{TS_END}"
        html = re.sub(re.escape(TS_START) + r'.*?' + re.escape(TS_END),
                      ts_block, html, flags=re.DOTALL)
        print(f"  ✓ Updated CENTRE_DATA_TIMESTAMPS: {ts_dict}")
    elif TS_START not in html:
        print(f"  ⚠ CENTRE_TIMESTAMPS markers not found — skipping")

    # 5. UPCOMING_DATA block
    if upcoming is not None and UP_START in html:
        ie_up, bsp_up, shoonya_up, samyama_up, hatha_up, all_up = upcoming
        fetched_str = datetime.datetime.now().strftime('%d %b %Y, %I:%M %p')
        up_block = (
            f"{UP_START}\n"
            f"// Upcoming programs pre-fetched from Isha API — last updated {fetched_str}\n"
            f"// The browser live fetch will override this data if it succeeds.\n"
            f"IE_UPCOMING      = {json.dumps(ie_up,      ensure_ascii=False)};\n"
            f"BSP_UPCOMING     = {json.dumps(bsp_up,     ensure_ascii=False)};\n"
            f"SHOONYA_UPCOMING = {json.dumps(shoonya_up, ensure_ascii=False)};\n"
            f"SAMYAMA_UPCOMING = {json.dumps(samyama_up, ensure_ascii=False)};\n"
            f"HATHA_UPCOMING   = {json.dumps(hatha_up,   ensure_ascii=False)};\n"
            f"ALL_UPCOMING     = {json.dumps(all_up,     ensure_ascii=False)};\n"
            f"/* UPCOMING_FETCHED: {fetched_str} */\n"
            f"{UP_END}"
        )
        html = re.sub(re.escape(UP_START) + r'.*?' + re.escape(UP_END),
                      up_block, html, flags=re.DOTALL)
        print(f"  ✓ Updated UPCOMING_DATA (fetched {fetched_str})")
    elif UP_START not in html:
        print(f"  ⚠ UPCOMING_DATA markers not found — skipping")

    # 6. IE_ONLINE_DATA block
    if ieo_data is not None and IEO_START in html:
        ieo_block = (f"{IEO_START}\nconst IE_ONLINE_DATA = "
                     f"{json.dumps(ieo_data, ensure_ascii=False)};\n{IEO_END}")
        html = re.sub(re.escape(IEO_START) + r'.*?' + re.escape(IEO_END),
                      ieo_block, html, flags=re.DOTALL)
        print(f"  ✓ Updated IE_ONLINE_DATA ({len(ieo_data)} centres)")
    elif IEO_START not in html:
        print(f"  ⚠ IE_ONLINE markers not found — skipping")

    # 7. IEO_CURRENT_MONTH block
    if ieo_cm_data is not None and IEO_CM_START in html:
        cm_block = (f"{IEO_CM_START}\nconst IEO_CURRENT_MONTH = "
                    f"{json.dumps(ieo_cm_data, ensure_ascii=False)};\n{IEO_CM_END}")
        html = re.sub(re.escape(IEO_CM_START) + r'.*?' + re.escape(IEO_CM_END),
                      cm_block, html, flags=re.DOTALL)
        print(f"  ✓ Updated IEO_CURRENT_MONTH ({ieo_cm_data.get('month','?')})")
    elif IEO_CM_START not in html:
        print(f"  ⚠ IEO_CURRENT_MONTH markers not found — skipping")

    with open(HTML, 'w', encoding='utf-8') as f:
        f.write(html)

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--date', help='Override history date key (YYYY-MM-DD), e.g. 2026-08-17')
    parser.add_argument('file', nargs='?', help='Optional single file path')
    args, _ = parser.parse_known_args()

    print("\n── Karnataka Registration Counts Updater ──")

    files = find_all_files()
    print(f"  Found {len(files)} file(s) in Santhosha Data:")
    for f in files:
        print(f"    • {os.path.relpath(f, SANTHOSHA_DIR)}")

    # Read current HTML once
    with open(HTML, 'r', encoding='utf-8') as f:
        html = f.read()

    # Read existing CENTRE_DATA and MONTHLY_DATA
    centre_data, _ = read_centre_data(html)
    if centre_data is None:
        print("  ⚠ Could not parse CENTRE_DATA from HTML — will skip Centre Insights update")

    monthly_data = read_monthly_data(html)
    if monthly_data is None:
        print("  ⚠ Could not parse MONTHLY_DATA from HTML — will skip monthly chart update")

    regs            = {}   # merged LIVE_REGS
    lang_map        = {}   # language per (centre, prog_key)
    cd_updates      = {}   # CENTRE_DATA changes to apply
    monthly_updates = {}   # MONTHLY_DATA changes to apply
    file_timestamps = {}   # per-centre XLS file mtimes (html_centre → mtime float)
    tmp_files       = []
    latest_mtime    = max(os.path.getmtime(f) for f in files)

    for xlsx in files:
        label = os.path.relpath(xlsx, SANTHOSHA_DIR)
        ext   = os.path.splitext(xlsx)[1].lower()

        # Try CRM "Total row" format first (xlsx only — no CSV export needed)
        if ext in ('.xlsx', '.xls') and HAS_OPENPYXL:
            if try_parse_crm_xlsx(xlsx, regs, cd_updates, monthly_updates, label, file_timestamps):
                continue   # handled — skip CSV export path

        # Standard batch-level format (Numbers files + non-CRM xlsx)
        if ext == '.numbers' and sys.platform != 'darwin':
            print(f"  ⚠ Skipping {label} (.numbers requires macOS)")
            continue
        tmp_csv = tempfile.mktemp(suffix=".csv")
        tmp_files.append(tmp_csv)
        print(f"\nExporting {label} → CSV …")
        if export_to_csv(xlsx, tmp_csv):
            parse_csv(tmp_csv, regs, cd_updates, label, lang_map=lang_map)
        else:
            print(f"  ⚠ Skipping {label} (export failed)")

    # Apply cd_updates → patch CENTRE_DATA
    if centre_data and cd_updates:
        changed = []
        for html_centre, progs in cd_updates.items():
            centre_data.setdefault(html_centre, {})
            for cd_key, year_counts in progs.items():
                centre_data[html_centre].setdefault(cd_key, {})
                for year, count in year_counts.items():
                    old = centre_data[html_centre][cd_key].get(year, 0)
                    centre_data[html_centre][cd_key][year] = count
                    if old != count:
                        changed.append(f"  {html_centre}/{cd_key}/{year}: {old} → {count}")
        if changed:
            print(f"\nCENTRE_DATA changes:")
            for c in changed: print(c)
        else:
            print(f"\nCENTRE_DATA: no changes")

    # Apply monthly_updates → patch MONTHLY_DATA
    if monthly_data and monthly_updates:
        changed = []
        for html_centre, progs in monthly_updates.items():
            monthly_data.setdefault(html_centre, {})
            for cd_key, year_months in progs.items():
                monthly_data[html_centre].setdefault(cd_key, {})
                for year, months in year_months.items():
                    monthly_data[html_centre][cd_key].setdefault(year, {})
                    for month, count in months.items():
                        old = monthly_data[html_centre][cd_key][year].get(month, 0)
                        monthly_data[html_centre][cd_key][year][month] = count
                        if old != count:
                            changed.append(f"  {html_centre}/{cd_key}/{year}-{month}: {old} → {count}")
        if changed:
            print(f"\nMONTHLY_DATA changes:")
            for c in changed: print(c)
        else:
            print(f"\nMONTHLY_DATA: no changes")

    # Print LIVE_REGS summary
    total = sum(sum(p.values()) for p in regs.values())
    print(f"\n{'─'*62}")
    print(f"  {'CENTRE':<22}  {'PROGRAMME':<28}  {'COUNT':>5}")
    print(f"{'─'*62}")
    for centre in sorted(regs):
        for prog_key, count in sorted(regs[centre].items()):
            if '|' in prog_key:
                prog, date = prog_key.split('|', 1)
                prog_display = f"{prog.title()} ({date})"
            else:
                prog_display = prog_key.title()
            print(f"  {centre:<22}  {prog_display:<28}  {count:>5}")
    print(f"{'─'*62}")
    print(f"  {'TOTAL':<52}  {total:>5}")

    # Fetch upcoming programs from Isha API
    print(f"\nFetching upcoming programs from Isha API …")
    upcoming = fetch_upcoming_programs()

    # Parse IE Online data — glob for any matching file (handles version suffixes like -3)
    import glob as _glob
    ieo_candidates = sorted(_glob.glob(os.path.join(SANTHOSHA_DIR, 'IE Online - All center*.xlsx')))
    ieo_data = None
    if ieo_candidates:
        ieo_path = ieo_candidates[-1]  # use latest (alphabetically last)
        print(f"\nParsing IE Online data from {os.path.basename(ieo_path)} …")
        ieo_data = parse_ie_online(ieo_path)
    else:
        print(f"\n  ⚠ IE Online file not found in {SANTHOSHA_DIR}")

    # Parse IE Online current month data
    cm_candidates = sorted(_glob.glob(os.path.join(SANTHOSHA_DIR, 'IE_Online_CurrentMonthStatus*.xlsx')))
    ieo_cm_data = None
    if cm_candidates:
        cm_path = cm_candidates[-1]
        print(f"\nParsing IE Online current month from {os.path.basename(cm_path)} …")
        ieo_cm_data = parse_ieo_current_month(cm_path)
        if ieo_cm_data is not None:
            cm_mtime = os.path.getmtime(cm_path)
            ieo_cm_data['as_of'] = datetime.datetime.fromtimestamp(cm_mtime).strftime('%d %b %Y')
    else:
        print(f"\n  ⚠ IE Online current month file not found in {SANTHOSHA_DIR}")

    # Build 3-day registration history — use Pivot Event file's mtime as the date
    pivot_files = [f for f in files if 'pivot event' in os.path.basename(f).lower()]
    if pivot_files:
        pivot_mtime = max(os.path.getmtime(f) for f in pivot_files)
        today_str = datetime.datetime.fromtimestamp(pivot_mtime).strftime('%Y-%m-%d')
    else:
        today_str = datetime.datetime.fromtimestamp(latest_mtime).strftime('%Y-%m-%d')
    regs_history = read_regs_history(html)
    # Remove existing entry for today (will be replaced with fresh data)
    regs_history = [h for h in regs_history if h.get('date') != today_str]
    # Append today's snapshot
    regs_history.append({'date': today_str, 'regs': regs})
    # Sort chronologically so slice(-3) in JS always picks the 3 most recent dates
    regs_history.sort(key=lambda h: h['date'])
    # Keep all history (display limit handled in the dashboard UI)

    print(f"\nInjecting into dashboard …")
    inject_html(
        html,
        regs,
        centre_data if cd_updates else None,
        monthly_data if monthly_updates else None,
        latest_mtime,
        file_timestamps if file_timestamps else None,
        upcoming,
        ieo_data,
        ieo_cm_data,
        regs_history,
        lang_map=lang_map,
    )

    # Cleanup
    for t in tmp_files:
        try: os.remove(t)
        except: pass

    print(f"\n✓ Done — {len(files)} file(s) processed.\n")
